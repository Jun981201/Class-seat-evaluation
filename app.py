"""Flask web application for classroom management tools.
Supports: seating chart (7-per-table), evaluation forms, duty schedule, auto-comments.
"""
import os
import re
import json
from datetime import datetime
from collections import OrderedDict

from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename

from generators import (
    SeatingGenerator, EvaluationGenerator, DutyGenerator, CommentGenerator
)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['TEMPLATES_FOLDER'] = os.path.join(os.path.dirname(__file__), 'evaluation_templates')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'output'), exist_ok=True)

ALLOWED_EXTENSIONS = {'xls', 'xlsx'}


def load_templates():
    """Load evaluation templates from templates.json."""
    templates_file = os.path.join(app.config['TEMPLATES_FOLDER'], 'templates.json')
    if os.path.exists(templates_file):
        with open(templates_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def find_template(template_id):
    """Find a template by id."""
    templates = load_templates()
    for t in templates:
        if t.get('id') == template_id:
            return t
    return None


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def parse_template_xlsx(filepath):
    """Auto-parse a .xlsx evaluation template and return a config dict.

    Detects:
      - Title from row 1
      - Column headers, widths, teacher initials, scoring rules
      - Group scoring section at the bottom
      - Merged cells for remark/special columns
    """
    import openpyxl
    from openpyxl.utils import get_column_letter as gcl

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    # ── Locate header row (must contain both 学号 and 姓名) ──
    header_row = None
    for r in range(1, min(max_row, 12) + 1):
        row_vals = [str(ws.cell(row=r, column=c).value or '').strip()
                    for c in range(1, min(max_col, 20) + 1)]
        if '学号' in row_vals and '姓名' in row_vals:
            header_row = r
            break
    if header_row is None:
        raise ValueError('未在模板中找到"学号"和"姓名"列，请确认模板格式')

    # ── Detect teacher / rule / info rows ──
    # Teacher row: short names like "王", "冯", "王、冯、孟"
    # Rule row: patterns like "1/0.6/0.2", "(-)", "+\\-", "1.5\\1\\0.5\\0"
    rule_pattern = re.compile(r'^[\d.]+[/\\]|^\(\D|^\+\\|^\.\d+-')

    teacher_row = header_row + 1
    rule_row = header_row + 2
    has_teacher = False
    has_rule = False

    for c in range(1, max_col + 1):
        tv = str(ws.cell(row=teacher_row, column=c).value or '').strip()
        rv = str(ws.cell(row=rule_row, column=c).value or '').strip()
        if tv and len(tv) <= 20 and not rule_pattern.match(tv):
            has_teacher = True
        if rv and rule_pattern.match(rv):
            has_rule = True

    # ── Read title from row 1 ──
    title = str(ws.cell(row=1, column=1).value or '').strip()
    if not title:
        # Try merged cells
        for mc in ws.merged_cells.ranges:
            if mc.min_row == 1:
                title = str(ws.cell(row=1, column=mc.min_col).value or '').strip()
                if title:
                    break

    # ── Read column definitions ──
    columns = []
    score_col_name = '总成绩'
    remark_col_name = '备注'
    score_col_width = 13.0
    remark_col_width = 13.0
    remark_col2_width = 13.0

    for c in range(1, max_col + 1):
        cl = gcl(c)
        width = ws.column_dimensions[cl].width
        if width is None:
            width = 8.0
        header_val = str(ws.cell(row=header_row, column=c).value or '').strip()

        if not header_val:
            # Check if this is part of a merged remark cell
            # If column has no header but previous was 备注, it's part of remark merge
            if c > 1 and remark_col_width != 13.0 and remark_col2_width == 13.0:
                remark_col2_width = width
            continue

        # Score / Remark columns
        if header_val == '总成绩':
            score_col_width = width
            continue
        if header_val in ('备注', '评语'):
            remark_col_name = header_val
            remark_col_width = width
            continue
        if header_val == '总评':
            remark_col_name = header_val
            remark_col_width = width
            continue

        # Regular column
        col_entry = {'header': header_val, 'width': width}
        col_entry['rule'] = ''
        col_entry['teacher'] = ''

        if has_teacher:
            tv = str(ws.cell(row=teacher_row, column=c).value or '').strip()
            col_entry['teacher'] = tv
        if has_rule:
            rv = str(ws.cell(row=rule_row, column=c).value or '').strip()
            col_entry['rule'] = rv

        columns.append(col_entry)

    # If remark_col2 wasn't detected, assume it's the next column after remark
    if remark_col2_width == 13.0 and remark_col_width != 13.0:
        # Check next column width
        for mc in ws.merged_cells.ranges:
            # Find merge that includes the remark column
            pass

    # ── Detect group scoring section ──
    group_section = None
    group_start = None
    for r in range(max(header_row + 3, max_row - 25), max_row + 1):
        for c in range(1, min(max_col, 5) + 1):
            v = str(ws.cell(row=r, column=c).value or '').strip()
            if v in ('小组分数', '组号'):
                group_start = r
                break
        if group_start:
            break

    if group_start:
        group_dims = []
        # Find the sub-header row (contains dimension names like 作品安装, 模型实现 etc.)
        # Look through rows starting from group_start to find non-identity, non-empty headers
        for sr in range(group_start, min(group_start + 5, max_row + 1)):
            dim_row_vals = []
            for c in range(2, max_col + 1):
                v = str(ws.cell(row=sr, column=c).value or '').strip()
                if v and v not in ('', '总分', '总成绩', '规则', '评分标准') and len(v) >= 2:
                    # Check this isn't a teacher row (short names)
                    has_teacher_like = all(len(x) <= 3 for x in v.replace('、', ' ').replace('，', ' ').split())
                    if not has_teacher_like or len(v) > 8:
                        dim_row_vals.append((c, v))
            if len(dim_row_vals) >= 2:
                # This looks like the dimension header row
                for c, v in dim_row_vals:
                    below1 = str(ws.cell(row=sr + 1, column=c).value or '').strip()
                    below2 = str(ws.cell(row=sr + 2, column=c).value or '').strip()
                    # Determine which is teacher vs rule
                    if rule_pattern.match(below1):
                        rule_v, teacher_v = below1, below2
                    elif rule_pattern.match(below2):
                        rule_v, teacher_v = below2, below1
                    else:
                        rule_v, teacher_v = below1, ''  # could be teacher only
                    group_dims.append({
                        'header': v,
                        'rule': rule_v if rule_v and rule_pattern.match(rule_v) else '',
                        'teacher': teacher_v if teacher_v and len(teacher_v) <= 15 else '',
                    })
                break

        # Count group data rows (numbered rows after headers)
        num_groups = 0
        data_start = sr + 2 if group_dims else group_start + 3
        for r in range(data_start, min(data_start + 20, max_row + 1)):
            gid = str(ws.cell(row=r, column=1).value or '').strip()
            try:
                if int(gid) >= 1:
                    num_groups = max(num_groups, int(gid))
            except ValueError:
                pass

        if group_dims:
            group_section = {
                'groups': group_dims,
                'num_groups': num_groups or 6,
            }

    # ── Detect title font/size ──
    title_font_name = '华文新魏'
    title_size = 18
    try:
        title_cell = ws.cell(row=1, column=1)
        if title_cell.font:
            if title_cell.font.name:
                title_font_name = title_cell.font.name
            if title_cell.font.size:
                title_size = title_cell.font.size
    except Exception:
        pass

    # ── Generate id from title ──
    # e.g. "黄浦劳技中心投石车项目过程性评价表" → "catapult" or auto-generated
    import hashlib
    tid = 'tpl_' + hashlib.md5(title.encode()).hexdigest()[:8]

    # Try to extract a meaningful name
    name = title.replace('黄浦劳技中心', '').replace('项目学生课堂评价记录单', '').replace('项目过程性评价表', '')
    if not name:
        name = title[:8]

    wb.close()

    return {
        'id': tid,
        'name': name,
        'title': title,
        'columns': columns,
        'score_col': score_col_name,
        'score_col_width': score_col_width,
        'remark_col': remark_col_name,
        'remark_col_width': remark_col_width,
        'remark_col2_width': remark_col2_width,
        'group_section': group_section,
        'title_font': title_font_name,
        'title_size': title_size,
    }


def parse_project_info(footer_text):
    """Parse project assignment string from footer.

    Input format:
      "通用技术503：8-30；通用技术502：1-7，31-37"
      "通用技术204：11-35；通用技术102：1-10"

    Returns:
      OrderedDict of {project_name: [student_id_strings]}
      e.g., {'通用技术503': ['8','9',...,'30'], '通用技术502': ['1','2',...,'37']}
    """
    projects = OrderedDict()
    if not footer_text:
        return projects

    # Normalize
    text = footer_text.replace('\n', ' ').replace('\r', ' ')

    # Find all "project_name：ranges" or "project_name:ranges" patterns
    # Project names are like "通用技术503", "通用技术102", etc.
    pattern = r'([^；;\s：:]*[术技]\d+)\s*[：:]\s*([^；;通]*)'
    matches = re.findall(pattern, text)

    for proj_name, range_str in matches:
        proj_name = proj_name.strip()
        range_str = range_str.strip()

        if not range_str:
            continue

        student_ids = []
        range_parts = re.split(r'[，,]\s*', range_str)
        for rp in range_parts:
            rp = rp.strip().rstrip('；;')
            if not rp:
                continue
            if '-' in rp:
                try:
                    start, end = rp.split('-', 1)
                    start, end = int(start.strip()), int(end.strip())
                    for sid in range(start, end + 1):
                        student_ids.append(str(sid).zfill(2))
                except ValueError:
                    try:
                        student_ids.append(str(int(rp.strip())).zfill(2))
                    except ValueError:
                        pass
            else:
                try:
                    student_ids.append(str(int(rp.strip())).zfill(2))
                except ValueError:
                    pass

        if student_ids:
            projects[proj_name] = student_ids

    return projects


def parse_weekly_student_list(filepath):
    """Parse the weekly student list Excel file.

    Returns:
        {
            'classes': {
                '高一1班': {
                    'students': [{id, name, gender}],
                    'projects': {'通用技术403': ['08','09',...], ...},
                    'footer_text': '...',
                }
            },
            'all_projects': ['通用技术403', '通用技术102', ...]
        }
    """
    import xlrd
    import openpyxl

    result = {'classes': OrderedDict(), 'all_projects': []}
    all_projects_set = set()

    if str(filepath).endswith('.xls'):
        wb = xlrd.open_workbook(filepath)
        sheet_names = wb.sheet_names()
        for sn in sheet_names:
            ws = wb.sheet_by_name(sn)
            students = []
            footer_text = ''
            header_row = None

            for r in range(ws.nrows):
                # Check if this is header row
                row_vals = [str(ws.cell_value(r, c)).strip() for c in range(min(ws.ncols, 6))]
                if '姓名' in row_vals and '学号' in row_vals:
                    header_row = r
                    continue

                if header_row is None:
                    continue

                # Check if this is a regular data row or footer
                first_val = str(ws.cell_value(r, 0)).strip()
                second_val = str(ws.cell_value(r, 1)).strip()

                # If first column doesn't contain grade info and second is not a number-like id,
                # it might be a footer
                if first_val and ('人' in first_val or '通用技术' in first_val or '班主任' in first_val):
                    # This is footer text
                    footer_text = ' '.join(str(ws.cell_value(r, c)).strip()
                                          for c in range(ws.ncols))
                    continue

                if not first_val and not second_val:
                    continue

                # Try to parse as student row
                sid = str(ws.cell_value(r, 1)).strip()  # column B is 学号
                name = str(ws.cell_value(r, 2)).strip()  # column C is 姓名
                gender = str(ws.cell_value(r, 3)).strip()  # column D is 性别

                # Validate: sid should be numeric
                if sid and name:
                    try:
                        # Pad to 2 digits
                        sid_int = int(sid)
                        sid = str(sid_int).zfill(2)
                        students.append({
                            'id': sid,
                            'name': name,
                            'gender': gender or '男',
                        })
                    except ValueError:
                        pass

            projects = parse_project_info(footer_text)
            for p in projects:
                all_projects_set.add(p)

            result['classes'][sn] = {
                'students': students,
                'projects': projects,
                'footer_text': footer_text,
                'student_count': len(students),
            }
    else:
        # .xlsx format
        wb = openpyxl.load_workbook(filepath)
        for sn in wb.sheetnames:
            ws = wb[sn]
            students = []
            footer_text = ''
            header_row = None

            for r in range(1, (ws.max_row or 1) + 1):
                row_vals = [str(ws.cell(row=r, column=c).value or '').strip()
                           for c in range(1, min((ws.max_column or 1), 7) + 1)]
                if '姓名' in row_vals and '学号' in row_vals:
                    header_row = r
                    continue
                if header_row is None:
                    continue

                first_val = str(ws.cell(row=r, column=1).value or '').strip()
                if first_val and ('人' in first_val or '通用技术' in first_val or '班主任' in first_val):
                    footer_text = ' '.join(
                        str(ws.cell(row=r, column=c).value or '').strip()
                        for c in range(1, min((ws.max_column or 1), 7) + 1))
                    continue

                sid = str(ws.cell(row=r, column=2).value or '').strip()
                name = str(ws.cell(row=r, column=3).value or '').strip()
                gender = str(ws.cell(row=r, column=4).value or '').strip()

                if sid and name:
                    try:
                        sid_int = int(sid)
                        sid = str(sid_int).zfill(2)
                        students.append({
                            'id': sid,
                            'name': name,
                            'gender': gender or '男',
                        })
                    except ValueError:
                        pass

            projects = parse_project_info(footer_text)
            for p in projects:
                all_projects_set.add(p)

            result['classes'][sn] = {
                'students': students,
                'projects': projects,
                'footer_text': footer_text,
                'student_count': len(students),
            }
        wb.close()

    result['all_projects'] = sorted(all_projects_set)
    return result


def filter_students_by_project(class_data, project_name):
    """Filter students in a class by project assignment."""
    if not project_name or project_name not in class_data.get('projects', {}):
        return class_data.get('students', [])

    project_ids = set(class_data['projects'][project_name])
    return [s for s in class_data['students'] if s['id'] in project_ids]


def parse_grades_file(filepath_or_bytes, is_xls=False):
    """Parse a grade import file and return students with gender and grade info.

    Accepts either a file path or raw bytes content. When passing bytes,
    is_xls must be set correctly to choose the parser.
    """
    import xlrd
    import openpyxl
    from io import BytesIO

    students = []
    metadata = {}

    content = filepath_or_bytes
    if isinstance(content, (str,)) and not is_xls:
        # It's a file path
        if str(content).endswith('.xls'):
            is_xls = True
        else:
            is_xls = False

    if is_xls:
        if isinstance(content, (str,)):
            wb = xlrd.open_workbook(content)
        else:
            wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)
        header_row = None
        for r in range(min(ws.nrows, 10)):
            if str(ws.cell_value(r, 0)).strip() == '学校':
                header_row = r
                break
        if header_row is None:
            raise ValueError('未找到表头行')

        for r in range(header_row + 1, ws.nrows):
            school = str(ws.cell_value(r, 0)).strip()
            grade = str(ws.cell_value(r, 1)).strip()
            class_name = str(ws.cell_value(r, 2)).strip()
            name = str(ws.cell_value(r, 3)).strip()
            gender = str(ws.cell_value(r, 4)).strip()
            sid = str(ws.cell_value(r, 5)).strip()
            grade_level = str(ws.cell_value(r, 6)).strip()
            keywords = str(ws.cell_value(r, 7)).strip() if ws.ncols > 7 else ''

            if not name or not sid:
                continue

            students.append({
                'school': school, 'grade': grade, 'class': class_name,
                'name': name, 'gender': gender, 'id': sid.zfill(2),
                'grade_level': grade_level, 'keywords': keywords, 'comment': '',
            })

        if students:
            metadata = {
                'school': students[0]['school'],
                'grade': students[0]['grade'],
                'class': students[0]['class'],
            }
    else:
        if isinstance(content, (str,)):
            wb = openpyxl.load_workbook(content)
        else:
            wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        header_row = None
        for r in range(1, min((ws.max_row or 1), 10) + 1):
            if str(ws.cell(row=r, column=1).value or '').strip() == '学校':
                header_row = r
                break
        if header_row is None:
            wb.close()
            raise ValueError('未找到表头行')

        for r in range(header_row + 1, (ws.max_row or 1) + 1):
            school = str(ws.cell(row=r, column=1).value or '').strip()
            grade = str(ws.cell(row=r, column=2).value or '').strip()
            class_name = str(ws.cell(row=r, column=3).value or '').strip()
            name = str(ws.cell(row=r, column=4).value or '').strip()
            gender = str(ws.cell(row=r, column=5).value or '').strip()
            sid = str(ws.cell(row=r, column=6).value or '').strip()
            grade_level = str(ws.cell(row=r, column=7).value or '').strip()
            keywords = str(ws.cell(row=r, column=8).value or '').strip()

            if not name or not sid:
                continue

            students.append({
                'school': school, 'grade': grade, 'class': class_name,
                'name': name, 'gender': gender, 'id': sid.zfill(2),
                'grade_level': grade_level, 'keywords': keywords, 'comment': '',
            })

        if students:
            metadata = {
                'school': students[0]['school'],
                'grade': students[0]['grade'],
                'class': students[0]['class'],
            }
        wb.close()

    return students, metadata


# ── Routes ──

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/parse-student-list', methods=['POST'])
def parse_student_list():
    """Upload weekly student list and return parsed classes/projects."""
    if 'file' not in request.files:
        return jsonify({'error': '请上传文件'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '请选择文件'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': '仅支持 .xls 或 .xlsx 格式'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        result = parse_weekly_student_list(filepath)
        os.remove(filepath)

        # Simplify for JSON: convert OrderedDict
        classes_info = {}
        for cn, cd in result['classes'].items():
            classes_info[cn] = {
                'student_count': cd['student_count'],
                'projects': list(cd['projects'].keys()),
                'project_details': {k: len(v) for k, v in cd['projects'].items()},
            }

        return jsonify({
            'classes': classes_info,
            'all_projects': result['all_projects'],
            '_data': {  # Full data for session use
                cn: {
                    'students': cd['students'],
                    'projects': cd['projects'],
                }
                for cn, cd in result['classes'].items()
            }
        })
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'error': f'解析失败: {str(e)}'}), 400


@app.route('/api/get-class-students', methods=['POST'])
def get_class_students():
    """Get filtered students for a class + project combination."""
    data = request.json
    class_data = data.get('class_data', {})
    project_name = data.get('project', '')

    students = filter_students_by_project(
        {'students': class_data.get('students', []),
         'projects': class_data.get('projects', {})},
        project_name
    )

    return jsonify({
        'students': students,
        'count': len(students),
        'males': sum(1 for s in students if s.get('gender') == '男'),
        'females': sum(1 for s in students if s.get('gender') == '女'),
    })


@app.route('/api/generate-seating', methods=['POST'])
def generate_seating():
    """Generate seating chart with 7-per-table layout."""
    data = request.json
    students = data.get('students', [])
    school_name = data.get('school_name', '')
    class_name = data.get('class_name', '')
    project_name = data.get('project_name', '')
    arrange_mode = data.get('arrange_mode', 'random')
    separate_gender = data.get('separate_gender', True)
    generate_duty = data.get('generate_duty', True)

    duty_schedule = None
    if generate_duty and students:
        duty_schedule = DutyGenerator.generate(students, group_size=5, mode='rotate')

    wb = SeatingGenerator.generate(
        students, school_name, class_name,
        duty_schedule, arrange_mode, separate_gender
    )

    output_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'output')
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = secure_filename(f'座位表_{school_name}_{class_name}_{timestamp}')
    output_path = os.path.join(output_dir, f'{safe_name}.xlsx')
    wb.save(output_path)

    return jsonify({
        'download_url': f'/api/download/{os.path.basename(output_path)}',
        'filename': f'座位表_{school_name}_{class_name}.xlsx',
    })


@app.route('/api/generate-evaluation', methods=['POST'])
def generate_evaluation():
    """Generate evaluation form for selected students."""
    data = request.json
    students = data.get('students', [])
    school_name = data.get('school_name', '')
    class_name = data.get('class_name', '')
    project_name = data.get('project_name', '')
    template_id = data.get('template_id', '3d')

    template_config = find_template(template_id)

    wb = EvaluationGenerator.generate(
        students, school_name, class_name,
        project_name=project_name,
        student_count=len(students),
        template_config=template_config,
    )

    output_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'output')
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    # Use timestamp + template_id for safe filename (Chinese chars stripped by secure_filename)
    template_id = data.get('template_id', 'default')
    safe_name = f'evaluation_{template_id}_{timestamp}'
    output_path = os.path.join(output_dir, f'{safe_name}.xlsx')
    wb.save(output_path)

    display_name = f'评价表_{school_name}_{class_name}_{timestamp}.xlsx'
    return jsonify({
        'download_url': f'/api/download/{os.path.basename(output_path)}',
        'filename': display_name,
    })


@app.route('/api/list-templates', methods=['GET'])
def list_templates():
    """Return available evaluation templates."""
    templates = load_templates()
    return jsonify([{
        'id': t.get('id', ''),
        'name': t.get('name', ''),
        'title': t.get('title', ''),
        'columns_count': len(t.get('columns', [])),
        'has_group': bool((t.get('group_section') or {}).get('groups')),
    } for t in templates])


@app.route('/api/upload-template', methods=['POST'])
def upload_template():
    """Upload a template .xlsx file.

    Modes:
      - action=preview (default): parse and return config without saving
      - action=save: save config to templates.json (requires JSON body)
    """
    if request.is_json:
        data = request.get_json(silent=True) or {}
        action = data.get('action', 'preview')
    else:
        action = request.form.get('action', 'preview')

    if action == 'preview':
        if 'file' not in request.files:
            return jsonify({'error': '请上传模板文件'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '请选择文件'}), 400
        if not file.filename.lower().endswith('.xlsx'):
            return jsonify({'error': '模板仅支持 .xlsx 格式'}), 400

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            config = parse_template_xlsx(filepath)
            # Keep file for save step — clean up old previews after
            preview_path = os.path.join(app.config['UPLOAD_FOLDER'], 'preview_' + filename)
            if os.path.exists(preview_path):
                os.remove(preview_path)
            os.rename(filepath, preview_path)
            config['_preview_file'] = preview_path
            return jsonify({'template': config})
        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'error': f'模板解析失败: {str(e)}'}), 400

    elif action == 'save':
        # Save the edited config + copy xlsx
        data = request.get_json() or {}
        config = data.get('config', {})
        preview_file = data.get('preview_file', '')
        if not config or not config.get('id'):
            return jsonify({'error': '缺少模板配置'}), 400

        templates_dir = app.config['TEMPLATES_FOLDER']
        os.makedirs(templates_dir, exist_ok=True)

        # Copy the preview xlsx file to templates folder
        import shutil
        dest_xlsx = os.path.join(templates_dir, f"{config['id']}.xlsx")
        if os.path.exists(preview_file):
            shutil.copy2(preview_file, dest_xlsx)
            os.remove(preview_file)

        # Update templates.json
        templates_file = os.path.join(templates_dir, 'templates.json')
        templates = []
        if os.path.exists(templates_file):
            with open(templates_file, 'r', encoding='utf-8') as f:
                templates = json.load(f)

        existing = [i for i, t in enumerate(templates) if t.get('id') == config['id']]
        # Clean up non-serializable fields
        clean_config = {k: v for k, v in config.items() if not k.startswith('_')}
        if existing:
            templates[existing[0]] = clean_config
        else:
            templates.append(clean_config)

        with open(templates_file, 'w', encoding='utf-8') as f:
            json.dump(templates, f, ensure_ascii=False, indent=2)

        name = clean_config.get('name', '')
        return jsonify({
            'template': {
                'id': clean_config['id'],
                'name': name,
                'title': clean_config['title'],
                'columns_count': len(clean_config.get('columns', [])),
                'has_group': bool(clean_config.get('group_section')),
            },
            'message': f'模板"{name}"已保存',
        })

    return jsonify({'error': '无效的 action 参数'}), 400


@app.route('/api/update-template', methods=['POST'])
def update_template():
    """Update an existing template's config (name, columns, etc.)."""
    data = request.get_json() or {}
    config = data.get('config', {})
    if not config or not config.get('id'):
        return jsonify({'error': '缺少模板配置'}), 400

    templates_dir = app.config['TEMPLATES_FOLDER']
    templates_file = os.path.join(templates_dir, 'templates.json')

    if not os.path.exists(templates_file):
        return jsonify({'error': '模板列表不存在'}), 404

    with open(templates_file, 'r', encoding='utf-8') as f:
        templates = json.load(f)

    clean_config = {k: v for k, v in config.items() if not k.startswith('_')}
    found = False
    for i, t in enumerate(templates):
        if t.get('id') == clean_config['id']:
            templates[i] = clean_config
            found = True
            break

    if not found:
        return jsonify({'error': '模板不存在'}), 404

    with open(templates_file, 'w', encoding='utf-8') as f:
        json.dump(templates, f, ensure_ascii=False, indent=2)

    name = clean_config.get('name', '')
    return jsonify({'message': f'模板"{name}"已更新'})


@app.route('/api/delete-template', methods=['POST'])
def delete_template():
    """Delete a template by id."""
    data = request.get_json() or {}
    template_id = data.get('id', '')
    if not template_id:
        return jsonify({'error': '缺少模板ID'}), 400

    templates_dir = app.config['TEMPLATES_FOLDER']
    templates_file = os.path.join(templates_dir, 'templates.json')
    if not os.path.exists(templates_file):
        return jsonify({'error': '模板列表不存在'}), 404

    with open(templates_file, 'r', encoding='utf-8') as f:
        templates = json.load(f)

    templates = [t for t in templates if t.get('id') != template_id]

    with open(templates_file, 'w', encoding='utf-8') as f:
        json.dump(templates, f, ensure_ascii=False, indent=2)

    # Also remove the .xlsx file
    xlsx_path = os.path.join(templates_dir, f'{template_id}.xlsx')
    if os.path.exists(xlsx_path):
        os.remove(xlsx_path)

    return jsonify({'message': '模板已删除'})


@app.route('/api/list-templates-full', methods=['GET'])
def list_templates_full():
    """Return all templates with full config (for editing)."""
    templates = load_templates()
    return jsonify(templates)


@app.route('/api/upload-grades', methods=['POST'])
def upload_grades():
    """Upload grades file, auto-generate comments, return downloadable Excel."""
    if 'file' not in request.files:
        return jsonify({'error': '请上传文件'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '请选择文件'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': '仅支持 .xls 或 .xlsx 格式'}), 400

    # Read file bytes into memory to avoid Windows file-lock issues
    file_bytes = file.read()
    is_xls = file.filename.lower().endswith('.xls')

    try:
        students, metadata = parse_grades_file(file_bytes, is_xls=is_xls)

        # Generate comments for each student
        for s in students:
            s['comment'] = CommentGenerator.generate_comment(
                s.get('grade_level', '良好'),
                name=s.get('name'),
                gender=s.get('gender'),
                keywords=s.get('keywords', ''),
            )

        # Generate Excel
        output_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'output')
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        school = metadata.get('school', '未知')
        class_name = metadata.get('class', '未知')
        safe_school = secure_filename(school) or 'school'
        safe_class = secure_filename(class_name) or 'class'
        output_path = os.path.join(output_dir,
                                   f'comments_{safe_school}_{safe_class}_{timestamp}.xlsx')
        CommentGenerator.export_comments_excel(students, output_path, metadata)

        return jsonify({
            'students': students,
            'count': len(students),
            'metadata': metadata,
            'download_url': f'/api/download/{os.path.basename(output_path)}',
            'filename': f'成绩评语_{school}_{class_name}.xlsx',
        })
    except Exception as e:
        return jsonify({'error': f'处理失败: {str(e)}'}), 400


@app.route('/api/download/<filename>')
def download_file(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'output',
                           secure_filename(filename))
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在或已过期'}), 404
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/preview-seating', methods=['POST'])
def preview_seating():
    """Preview seating for 7-per-table layout."""
    data = request.json
    students = data.get('students', [])
    arrange_mode = data.get('arrange_mode', 'random')
    separate_gender = data.get('separate_gender', True)

    arranged = SeatingGenerator._shuffle_students(students, arrange_mode, separate_gender)

    positions = SeatingGenerator._assign_positions(arranged)

    return jsonify({
        'positions': positions,
        'total_seats': len(positions),
        'unassigned': [s for s in arranged[len(positions):]] if len(arranged) > len(positions) else [],
    })


if __name__ == '__main__':
    app.run(debug=True, port=5051)
