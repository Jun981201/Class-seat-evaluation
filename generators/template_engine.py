"""Core template engine - provides common Excel formatting utilities."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy

# ── Font definitions ──
FONT_TITLE = Font(name='宋体', size=18, bold=True)
FONT_TITLE_XINWEI = Font(name='华文新魏', size=18, bold=True)
FONT_INFO_LABEL = Font(name='黑体', size=10, bold=True)
FONT_INFO_VALUE = Font(name='黑体', size=10)
FONT_HEADER = Font(name='宋体', size=10, bold=True)
FONT_DATA = Font(name='宋体', size=11)
FONT_LEGEND = Font(name='宋体', size=10)

# ── Alignment definitions ──
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_CENTER_NOWRAP = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# ── Border definitions ──
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
BORDER_MEDIUM = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)
BORDER_BOTTOM_MEDIUM = Border(bottom=Side(style='medium'))
BORDER_ALL_MEDIUM = BORDER_MEDIUM

# ── Fill definitions ──
FILL_FEMALE = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
FILL_MALE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_HEADER_LIGHT = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')


class TemplateEngine:
    """Base class for Excel generation with common formatting methods."""

    @staticmethod
    def apply_cell(cell, value=None, font=None, alignment=None, border=None, fill=None):
        if value is not None:
            cell.value = value
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if fill:
            cell.fill = fill

    @staticmethod
    def apply_row(ws, row, col_start, col_end, font=None, alignment=None, border=None):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=c)
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border

    @staticmethod
    def set_col_widths(ws, widths_dict):
        for col_letter, width in widths_dict.items():
            ws.column_dimensions[col_letter].width = width

    @staticmethod
    def set_row_heights(ws, heights_dict):
        for row_num, height in heights_dict.items():
            ws.row_dimensions[row_num].height = height

    @staticmethod
    def merge_and_write(ws, merge_range, value, font=None, alignment=None, border=None, fill=None):
        ws.merge_cells(merge_range)
        top_left = merge_range.split(':')[0]
        cell = ws[top_left]
        if value is not None:
            cell.value = value
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            _apply_border_to_range(ws, merge_range, border)
        if fill:
            _apply_fill_to_range(ws, merge_range, fill)


def _apply_border_to_range(ws, cell_range, border):
    """Apply border to all cells in a range."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for r in range(int(min_row), int(max_row) + 1):
        for c in range(int(min_col), int(max_col) + 1):
            ws.cell(row=r, column=c).border = border


def _apply_fill_to_range(ws, cell_range, fill):
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for r in range(int(min_row), int(max_row) + 1):
        for c in range(int(min_col), int(max_col) + 1):
            ws.cell(row=r, column=c).fill = fill
