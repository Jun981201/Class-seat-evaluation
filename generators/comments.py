"""Auto-generate evaluation comments based on grade levels and optional keywords."""
import random

# Comment templates by grade level
COMMENT_TEMPLATES = {
    '优秀': [
        '学习态度端正，课堂表现突出，善于发现并解决技术问题。',
        '在3D建模和小组活动中表现优异，能独立完成复杂建模任务，作品完成度高，富有创意。胆大心细，勇于尝试，可以挑战更有难度的创意设计来突破自我。',
        '上课认真积极，善于思考和创新，小组贡献突出，承担了大部分的编程、建模和团队核心工作。始终保持积极的学习态度。',
        '学习效率高，课堂表现活跃，善于思考。乐于提问和探究新知识，富有创新精神，小组贡献突出。',
        '态度认真，能快速高质量完成任务，表现一直很稳定。有很强的工程思维，善于帮助小组同学解决实际问题。',
        '学习态度端正，课堂表现活跃，善于创新，善于思考与分析问题，是小组活动的核心骨干。',
        '动手能力强，建模速度快，有很强的学习能力。课堂表现积极，希望多在团队合作中发挥引领作用。',
        '设计思维出色，作品创意十足，善于将想法转化为实物。在团队中发挥关键领导作用，善于分配任务和协调。',
        '对新知识有强烈的学习兴趣，能快速掌握复杂技能。课堂上积极发言，带动全班学习氛围。',
        '学习态度端正，上课比较专注，希望多为团队作品做出贡献。',
        '积极参与小组活动，善于发现并解决技术问题，乐于帮助同学，综合表现都很好。',
        '学习态度端正，有较强的理解能力，希望在个人建模技术上继续突破，更加大胆地尝试。',
    ],
    '良好': [
        '学习态度较认真，能按时完成各项任务，课堂表现较好，技术水平较稳定。',
        '学习态度较好，设计构思有创意，能在规定时间完成项目。课堂纪律方面需要加强。',
        '上课认真听讲，遵守课堂纪律，学习态度较好，但在小组设计构思方面创意不足。',
        '能主动参与小组活动，学习态度端正，按时完成各项学习任务。整体表现良好。',
        '善于动手实践，热爱技术探索。课堂上善于观察思考，善于拓展解题思路。',
        '有好奇心，善于动手实践和解决问题，课堂表现需要更积极，偶尔会开小差。',
        '学习态度比较好，设计构思有创意，胜任作品制作任务。课堂纪律方面需要加强。',
        '学习态度端正，课堂表现需要更加积极，能领导小组活动。希望更加大胆地尝试。',
        '基本能掌握所学技能，工具使用规范。课堂表现较好，希望多参与动手实践。',
        '能参与小组活动，学习态度比较认真，有一定的合作意识。希望多动手实践。',
        '学习态度较好，上课认真专注，能按时完成各项学习任务。希望多参与讨论。',
        '对待学习认真细致，小组活动中能积极参与并主动与老师沟通。作为组长认真负责。',
    ],
    '合格': [
        '能完成老师布置的学习任务，但对学习投入度不足，对待学习较为被动、懒散。',
        '上课注意力不够集中，学习投入一般，主动性和合作意识亟需加强。希望端正学习态度，加以改进。',
        '上课偶有开小差状况，能完成小组分配的工作。建模速度快，期待能向老师展示更优秀的一面。',
        '上课态度比较浮躁，但能按时完成老师布置的任务。动手实践能力需要进一步加强。',
        '课堂投入不够，偶尔会做与学习无关的事。希望端正学习态度，提高学习投入度。',
        '能完成基本任务，但缺乏主动探索精神。希望多向优秀同学学习，提高自己。',
        '上课偶尔开小差，学习态度不够稳定。但有一定潜力，需要更多自律。',
        '对待学习比较被动，需要老师督促。希望能提高主动性，积极参与课堂活动。',
        '基本能跟上教学进度，但深度不够。希望加强课后练习，巩固所学知识。',
        '能在小组中完成分配的任务，但缺乏主动性。希望下次能承担更多责任。',
    ],
    '需努力': [
        '学习态度不够端正，多次违反课堂纪律，上课常睡觉或玩手机，不参与小组学习活动。需认真反思学习态度。',
        '课堂纪律意识差，多次被老师反映干扰学习活动，破坏学习工具和公物，行为举止较差。需改正不良习惯。',
        '上课投入度不够，对学习任务完成度低，课堂参与度不足。反复提醒无效，希望深刻反思并端正态度。',
        '课堂参与度低，学习投入明显不足，多次未完成课堂学习任务。纪律意识淡薄，需加强自我约束。',
        '学习态度不端正，上课经常随意离开座位，影响其他同学学习。经老师多次劝导效果甚微，需要深刻改正。',
        '缺乏学习动力，经常不做值日，作业不交。希望尽快调整状态，认真对待每一堂课。',
        '上课精神不集中，多次睡觉、玩手机等不良行为。小组协作能力差，任务完成度低。需深刻反思。',
    ],
}

# Keyword integration patterns — {keyword} is replaced with the actual keyword
KEYWORD_PHRASES = {
    '优秀': [
        '在{keyword}方面表现突出，能力出色',
        '在{keyword}方面有很强的创造力和执行力',
        '在{keyword}方面起到了示范引领作用',
        '在{keyword}上表现优异，值得表扬',
        '{keyword}能力突出，是同学们的榜样',
        '在{keyword}方面展现了出色的天赋和努力',
    ],
    '良好': [
        '在{keyword}方面表现较好，有一定基础',
        '在{keyword}方面能按时完成任务',
        '{keyword}方面有进步，继续保持',
        '在{keyword}上表现稳定，希望继续提升',
        '在{keyword}方面基本达标，还有提升空间',
    ],
    '合格': [
        '需要加强{keyword}方面的投入与练习',
        '在{keyword}方面需要更加主动和坚持',
        '{keyword}方面的表现还有较大提升空间',
        '希望在{keyword}上投入更多精力',
        '在{keyword}方面需端正态度，认真对待',
    ],
    '需努力': [
        '在{keyword}方面表现较差，亟需改进',
        '{keyword}方面多次未达标，需要严格要求自己',
        '在{keyword}方面缺乏基本投入，需深刻反思',
        '亟需在{keyword}方面做出实质性改变',
    ],
}

class CommentGenerator:
    """Generate evaluation comments based on grade level and optional keywords."""

    @staticmethod
    def generate_comment(grade, name=None, gender=None, keywords=None):
        """
        Generate a comment based on grade level and optional keywords.

        Args:
            grade: '优秀', '良好', '合格', or '需努力'
            name: optional student name
            gender: optional, for pronoun selection
            keywords: optional comma/space-separated keywords from template

        Returns:
            str: generated comment
        """
        kw = (keywords or '').strip()

        if kw:
            return CommentGenerator._generate_keyword_comment(grade, kw)
        else:
            templates = COMMENT_TEMPLATES.get(grade, COMMENT_TEMPLATES['良好'])
            return random.choice(templates)

    @staticmethod
    def _generate_keyword_comment(grade, keywords):
        """Generate a comment that incorporates the provided keywords."""
        import re
        kws = re.split(r'[,，、\s;；]+', keywords)
        kws = [k.strip() for k in kws if k.strip()]
        if not kws:
            templates = COMMENT_TEMPLATES.get(grade, COMMENT_TEMPLATES['良好'])
            return random.choice(templates)

        # Use at most 3 keywords to keep the comment concise
        selected = kws[:3]
        joined_kws = '、'.join(selected)

        # Pick a keyword phrase for the grade level
        phrases = KEYWORD_PHRASES.get(grade, KEYWORD_PHRASES['良好'])
        kw_phrase = random.choice(phrases).format(keyword=joined_kws)

        # Pick a random grade-level comment as the second sentence
        templates = COMMENT_TEMPLATES.get(grade, COMMENT_TEMPLATES['良好'])
        grade_comment = random.choice(templates)

        return f'{kw_phrase}。{grade_comment}'

    @classmethod
    def generate_for_students(cls, students, reference_file=None):
        """
        Generate comments for a list of students with grades and optional keywords.
        """
        result = []
        for s in students:
            grade = s.get('grade', '良好')
            comment = cls.generate_comment(
                grade,
                name=s.get('name'),
                gender=s.get('gender'),
                keywords=s.get('keywords', ''),
            )
            s_copy = dict(s)
            s_copy['comment'] = comment
            result.append(s_copy)
        return result

    @classmethod
    def generate_from_grades_file(cls, grades_file_path):
        """
        Read a grades import file and generate comments.
        Columns: 学校, 年级, 班级, 姓名, 性别, 学号, 成绩, 评语
        """
        if str(grades_file_path).endswith('.xls'):
            import xlrd
            wb = xlrd.open_workbook(grades_file_path)
            ws = wb.sheet_by_index(0)
            students = []
            metadata = {}

            header_row = None
            for r in range(min(ws.nrows, 10)):
                if str(ws.cell_value(r, 0)).strip() == '学校':
                    header_row = r
                    break

            if header_row is None:
                raise ValueError('未找到表头行（含"学校"列）')

            for r in range(header_row + 1, ws.nrows):
                school = str(ws.cell_value(r, 0)).strip()
                grade = str(ws.cell_value(r, 1)).strip()
                class_name = str(ws.cell_value(r, 2)).strip()
                name = str(ws.cell_value(r, 3)).strip()
                gender = str(ws.cell_value(r, 4)).strip()
                student_id = str(ws.cell_value(r, 5)).strip()
                grade_level = str(ws.cell_value(r, 6)).strip()
                keywords = str(ws.cell_value(r, 7)).strip() if ws.ncols > 7 else ''

                if not name or not student_id:
                    continue

                comment = cls.generate_comment(grade_level, name=name, gender=gender, keywords=keywords)
                students.append({
                    'school': school,
                    'grade': grade,
                    'class': class_name,
                    'name': name,
                    'gender': gender,
                    'id': student_id,
                    'grade_level': grade_level,
                    'keywords': keywords,
                    'comment': comment,
                })

            if not metadata and students:
                metadata = {
                    'school': students[0]['school'],
                    'grade': students[0]['grade'],
                    'class': students[0]['class'],
                }

            return students, metadata

        else:
            import openpyxl
            wb = openpyxl.load_workbook(grades_file_path)
            ws = wb.active
            students = []
            metadata = {}

            header_row = None
            for r in range(1, min(ws.max_row or 1, 10) + 1):
                if str(ws.cell(row=r, column=1).value or '').strip() == '学校':
                    header_row = r
                    break

            if header_row is None:
                raise ValueError('未找到表头行（含"学校"列）')

            for r in range(header_row + 1, (ws.max_row or 1) + 1):
                school = str(ws.cell(row=r, column=1).value or '').strip()
                grade = str(ws.cell(row=r, column=2).value or '').strip()
                class_name = str(ws.cell(row=r, column=3).value or '').strip()
                name = str(ws.cell(row=r, column=4).value or '').strip()
                gender = str(ws.cell(row=r, column=5).value or '').strip()
                student_id = str(ws.cell(row=r, column=6).value or '').strip()
                grade_level = str(ws.cell(row=r, column=7).value or '').strip()
                keywords = str(ws.cell(row=r, column=8).value or '').strip()

                if not name or not student_id:
                    continue

                comment = cls.generate_comment(grade_level, name=name, gender=gender, keywords=keywords)
                students.append({
                    'school': school,
                    'grade': grade,
                    'class': class_name,
                    'name': name,
                    'gender': gender,
                    'id': student_id,
                    'grade_level': grade_level,
                    'keywords': keywords,
                    'comment': comment,
                })

            if not metadata and students:
                metadata = {
                    'school': students[0]['school'],
                    'grade': students[0]['grade'],
                    'class': students[0]['class'],
                }

            return students, metadata

    @classmethod
    def export_comments_excel(cls, students, output_path, metadata=None):
        """
        Export students with comments to the 成绩导入模板 Excel format.
        """
        import openpyxl
        from .template_engine import (
            FONT_HEADER, FONT_DATA, ALIGN_CENTER, BORDER_THIN
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '成绩导入'

        for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'],
                          [14, 8, 8, 10, 6, 10, 10, 45]):
            ws.column_dimensions[col].width = w

        ws.cell(row=1, column=1, value='成绩表导入模板')
        ws.cell(row=2, column=1, value='1、请勿更改表格样式')
        ws.cell(row=3, column=1, value='2、所有列均为必填项，确保填写的信息与学生信息一致（班级、姓名、学号必须相同）')
        ws.cell(row=4, column=1, value='3、成绩栏可填写项为："优秀"、"良好"、"合格"、"需努力"')
        ws.cell(row=5, column=1, value='4、评语栏请将字数控制在100字以内，且尽量不使用换行')

        headers = ['学校', '年级', '班级', '姓名', '性别', '学号', '成绩', '评语']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=ci, value=h)
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN

        for si, s in enumerate(students):
            row = si + 8
            meta = metadata or {}
            values = [
                s.get('school', meta.get('school', '')),
                s.get('grade', meta.get('grade', '')),
                s.get('class', meta.get('class', '')),
                s.get('name', ''),
                s.get('gender', ''),
                s.get('id', ''),
                s.get('grade_level', ''),
                s.get('comment', ''),
            ]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = FONT_DATA
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_THIN

        wb.save(output_path)
        return output_path
