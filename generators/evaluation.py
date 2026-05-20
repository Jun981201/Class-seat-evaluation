"""Evaluation/scoring form generator — template-file-driven.

When a .xlsx template file exists, opens it as base and fills in student data,
preserving ALL original formatting exactly. Falls back to code-based generation
only when no template file is available.
"""
import os
from copy import copy

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .template_engine import (
    TemplateEngine, ALIGN_CENTER, BORDER_THIN, BORDER_BOTTOM_MEDIUM,
    FONT_INFO_LABEL, FONT_INFO_VALUE, FONT_HEADER, FONT_DATA,
)


# Template folder relative to project root
_TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                              'evaluation_templates')


class EvaluationGenerator(TemplateEngine):
    """Generate student evaluation record Excel sheets from template files."""

    @classmethod
    def generate(cls, students, school_name, class_name, project_name='',
                 student_count=None, template_config=None):
        if template_config is None:
            template_config = cls._default_config()

        n = student_count or len(students)
        template_id = template_config.get('id', '')
        template_xlsx = os.path.join(_TEMPLATES_DIR, f'{template_id}.xlsx')

        if os.path.exists(template_xlsx):
            return cls._generate_from_file(template_xlsx, students,
                                           school_name, class_name, n)

        # Fallback: no template file, build from JSON config
        return cls._generate_from_config(students, school_name, class_name,
                                         project_name, n, template_config)

    # ── Template-file-based generation ──

    @classmethod
    def _generate_from_file(cls, template_path, students, school_name,
                            class_name, student_count):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1

        # ---- Locate structure ----
        header_row = None
        for r in range(1, min(max_row, 15) + 1):
            row_vals = [str(ws.cell(row=r, column=c).value or '').strip()
                        for c in range(1, min(max_col, 5) + 1)]
            if '学号' in row_vals and '姓名' in row_vals:
                header_row = r
                break
        if header_row is None:
            raise ValueError('模板中未找到"学号"和"姓名"列')

        # Teacher row = next row if cells contain short Chinese-like names
        teacher_row = header_row + 1
        has_teacher = False
        for c in range(4, max_col + 1):
            tv = str(ws.cell(row=teacher_row, column=c).value or '').strip()
            if tv and len(tv) <= 15 and not any(ch in tv for ch in '/.\\-+('):
                has_teacher = True
                break

        # Rule row = next row if cells contain score patterns like 1/0.6/0.2
        rule_row = teacher_row + 1 if has_teacher else header_row + 1
        has_rule = False
        for c in range(4, max_col + 1):
            rv = str(ws.cell(row=rule_row, column=c).value or '').strip()
            if rv and any(ch in rv for ch in '/.\\'):
                has_rule = True
                break

        last_header_row = rule_row if has_rule else (teacher_row if has_teacher else header_row)
        data_start_row = last_header_row + 1

        # Find group section
        group_start_row = None
        for r in range(data_start_row, max_row + 1):
            for c in range(1, min(max_col + 1, 6)):
                v = str(ws.cell(row=r, column=c).value or '').strip()
                if v in ('小组分数',):
                    group_start_row = r
                    break
            if group_start_row:
                break

        if group_start_row is None:
            for r in range(data_start_row + 5, max_row + 1):
                v = str(ws.cell(row=r, column=1).value or '').strip()
                if v == '组号':
                    group_start_row = r - 1
                    break

        if group_start_row:
            data_end_row = group_start_row - 1
        else:
            data_end_row = max_row

        template_data_rows = data_end_row - data_start_row + 1
        needed_rows = student_count

        if needed_rows > template_data_rows:
            rows_to_add = needed_rows - template_data_rows
            insert_at = data_end_row + 1
            cls._insert_data_rows(ws, insert_at, rows_to_add,
                                  data_start_row, max_col)
            data_end_row += rows_to_add
            if group_start_row:
                group_start_row += rows_to_add

        elif needed_rows < template_data_rows:
            # Clear unused rows — keep formatting, only blank values.
            # MergedCell objects are read-only, so skip them.
            clear_start = data_start_row + needed_rows
            for r in range(clear_start, data_end_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.__class__.__name__ != 'MergedCell':
                        cell.value = None
            data_end_row = clear_start - 1

        # Fill info row
        info_row = None
        for r in range(1, header_row):
            for c in range(1, min(max_col + 1, 3)):
                if str(ws.cell(row=r, column=c).value or '').strip() == '学校':
                    info_row = r
                    break
            if info_row:
                break

        if info_row:
            cls._fill_info_row(ws, info_row, max_col, school_name, class_name, student_count)

        # Fill student data
        ref_row = data_start_row
        for si, student in enumerate(students):
            if si >= needed_rows:
                break
            row = data_start_row + si
            cls._ensure_row_format(ws, row, ref_row, max_col)
            cls._write_cell(ws, row, 1, student.get('id', ''))
            cls._write_cell(ws, row, 2, student.get('name', ''))
            cls._write_cell(ws, row, 3, student.get('gender', ''))

        return wb

    @staticmethod
    def _insert_data_rows(ws, insert_at, count, ref_row, max_col):
        """Insert rows while preserving formatting, row heights, and merged cells.

        openpyxl does NOT shift row_dimensions or merged_cells when rows are
        inserted. We must save them, do the insert, then re-apply with offsets.
        """
        # 1. Save original row heights for rows at/after insert point
        orig_heights = {}
        for r in range(insert_at, ws.max_row + 51):
            h = ws.row_dimensions[r].height
            if h is not None:
                orig_heights[r] = h

        # 2. Save and unmerge all merged cell ranges
        saved_merges = [str(m) for m in ws.merged_cells.ranges]
        for m in saved_merges:
            ws.unmerge_cells(m)

        # 3. Insert rows
        for i in range(count):
            ws.insert_rows(insert_at + i)
            EvaluationGenerator._copy_row_format(ws, ref_row, insert_at + i, max_col)

        # 4. Re-apply row heights with offsets
        for r in range(insert_at, ws.max_row + 51):
            ws.row_dimensions[r].height = None
        for old_row, h in orig_heights.items():
            ws.row_dimensions[old_row + count].height = h

        # 5. Re-apply merged cells with offsets
        for merge_str in saved_merges:
            new_range = EvaluationGenerator._shift_merge_range(
                merge_str, insert_at, count)
            if new_range:
                ws.merge_cells(new_range)

    @staticmethod
    def _shift_merge_range(merge_str, insert_at, offset):
        """Shift a merged cell range string past an insert point.

        If the range is entirely below the insert point, shift it down by offset.
        If entirely above, leave unchanged. If spanning, this is unexpected.
        """
        parts = merge_str.split(':')
        if len(parts) != 2:
            return merge_str

        r1, c1 = EvaluationGenerator._parse_cell_ref(parts[0])
        r2, c2 = EvaluationGenerator._parse_cell_ref(parts[1])

        if r1 >= insert_at:
            r1 += offset
            r2 += offset
        elif r2 >= insert_at:
            # Range spans the insert point — unexpected, but shift the part below
            r2 += offset

        from openpyxl.utils import get_column_letter
        return (f'{get_column_letter(c1)}{r1}:'
                f'{get_column_letter(c2)}{r2}')

    @staticmethod
    def _parse_cell_ref(ref):
        """Parse 'A1' -> (row, col)."""
        import re
        m = re.match(r'^([A-Z]+)(\d+)$', ref)
        if not m:
            return (0, 0)
        col_str = m.group(1)
        col = 0
        for ch in col_str:
            col = col * 26 + (ord(ch) - ord('A') + 1)
        return (int(m.group(2)), col)

    @staticmethod
    def _copy_row_format(ws, src_row, dst_row, max_col):
        """Copy cell formatting from src_row to dst_row."""
        for c in range(1, max_col + 1):
            src = ws.cell(row=src_row, column=c)
            dst = ws.cell(row=dst_row, column=c)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = src.number_format
                dst.alignment = copy(src.alignment)
            # Copy row height
            if ws.row_dimensions[src_row].height:
                ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    @staticmethod
    def _ensure_row_format(ws, row, ref_row, max_col):
        """Ensure the row has proper formatting from reference row."""
        # Check if row already has borders (from insert)
        ref = ws.cell(row=ref_row, column=1)
        cell = ws.cell(row=row, column=1)
        if cell.border and cell.border.left and cell.border.left.style:
            return  # Already formatted
        for c in range(1, max_col + 1):
            src = ws.cell(row=ref_row, column=c)
            dst = ws.cell(row=row, column=c)
            if src.has_style:
                dst.border = copy(src.border)
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.fill = copy(src.fill)

    @staticmethod
    def _write_cell(ws, row, col, value):
        """Write value to cell without changing existing formatting."""
        cell = ws.cell(row=row, column=col)
        cell.value = value

    @staticmethod
    def _fill_info_row(ws, info_row, max_col, school_name, class_name, student_count):
        """Fill school/class/count in info row."""
        for c in range(1, max_col + 1):
            label = str(ws.cell(row=info_row, column=c).value or '').strip()
            # Find the value cell next to each label
            if label == '学校':
                # Value is in the merged cell to the right
                ws.cell(row=info_row, column=c + 1).value = school_name
            elif label == '班级':
                ws.cell(row=info_row, column=c + 1).value = class_name
            elif label == '人数':
                ws.cell(row=info_row, column=c + 1).value = student_count

    # ── Fallback: code-based generation from JSON config ──

    @classmethod
    def _generate_from_config(cls, students, school_name, class_name,
                              project_name, student_count, config):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet('教学评表')
        cls._build_from_config(ws, students, school_name, class_name,
                               project_name, student_count, config)
        return wb

    @classmethod
    def _default_config(cls):
        return {
            'id': '3d', 'name': '3D建模项目',
            'title': '黄浦劳技中心3D项目学生课堂评价记录单',
            'columns': [
                {'header': '学号', 'width': 5.0},
                {'header': '姓名', 'width': 8.5},
                {'header': '性别', 'width': 6.0},
                {'header': '工具收纳', 'width': 5.0, 'rule': '(-)', 'teacher': ''},
                {'header': '切片', 'width': 8.5, 'rule': '1/.6/.2', 'teacher': ''},
                {'header': '三维建模\n特征/扫掠', 'width': 5.0, 'rule': '', 'teacher': ''},
                {'header': '团队协作', 'width': 5.0, 'rule': '+\\-', 'teacher': ''},
                {'header': '操作规范', 'width': 9.0, 'rule': '(-)', 'teacher': ''},
                {'header': '建模打印', 'width': 8.0, 'rule': '.5-1.5+\\-', 'teacher': ''},
                {'header': '小组作品总分', 'width': 9.5, 'rule': '', 'teacher': ''},
                {'header': '完整度', 'width': 5.5, 'rule': '', 'teacher': ''},
                {'header': '个人作品', 'width': 5.5, 'rule': '', 'teacher': ''},
                {'header': '总分', 'width': 7.5, 'rule': '', 'teacher': ''},
            ],
            'score_col': '总成绩', 'remark_col': '备注',
            'score_col_width': 13, 'remark_col_width': 13, 'remark_col2_width': 13,
            'group_section': {
                'groups': [
                    {'header': '模型实现', 'rule': '1/.6/.2'},
                    {'header': '作品包装', 'rule': '1/.6/.2'},
                    {'header': '功能测试', 'rule': '1/.6/.2'},
                    {'header': '汇报表现', 'rule': '1/.6/.2'},
                ],
                'num_groups': 8,
            },
            'title_font': '华文新魏', 'title_size': 18,
        }

    @classmethod
    def _build_from_config(cls, ws, students, school_name, class_name,
                           project_name, student_count, config):
        columns = config['columns']
        score_col_name = config.get('score_col', '总成绩')
        remark_col_name = config.get('remark_col', '备注')
        group_cfg = config.get('group_section', {})
        group_dims = group_cfg.get('groups', [])
        num_groups = group_cfg.get('num_groups', 8)

        ident_cols = columns[:3]
        dim_cols = columns[3:]
        num_dims = len(dim_cols)
        total_cols = 3 + num_dims + 1 + 2

        has_teacher = any(c.get('teacher', '') for c in dim_cols)
        has_rule = any(c.get('rule', '') for c in dim_cols)
        multi_row_header = has_teacher or has_rule

        # Column widths
        widths = {}
        for i, col in enumerate(columns):
            widths[get_column_letter(i + 1)] = col.get('width', 8)
        score_col_idx = len(columns) + 1
        remark_col_idx = score_col_idx + 1
        widths[get_column_letter(score_col_idx)] = config.get('score_col_width', 13)
        widths[get_column_letter(remark_col_idx)] = config.get('remark_col_width', 13)
        widths[get_column_letter(remark_col_idx + 1)] = config.get('remark_col2_width', 13)
        cls.set_col_widths(ws, widths)

        # Row heights
        ws.row_dimensions[1].height = 21
        ws.row_dimensions[2].height = 3
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 4.5
        ws.row_dimensions[5].height = 26 if multi_row_header else 23.25
        if has_teacher:
            ws.row_dimensions[6].height = 15
        if has_rule:
            ws.row_dimensions[7].height = 10

        last_col = get_column_letter(total_cols)

        # Title
        title_font = Font(name=config.get('title_font', '华文新魏'),
                          size=config.get('title_size', 18), bold=True)
        cls.merge_and_write(ws, f'A1:{last_col}1', config['title'],
                           title_font, ALIGN_CENTER)

        # Info row
        cls._build_info_row(ws, school_name, class_name, student_count,
                           total_cols, score_col_idx)

        # Headers
        if multi_row_header:
            cls._build_multi_row_header(ws, columns, dim_cols,
                                        score_col_name, remark_col_name,
                                        total_cols, has_teacher, has_rule)
            data_start_row = 8
        else:
            cls._build_single_row_header(ws, columns, dim_cols,
                                         score_col_name, remark_col_name,
                                         total_cols)
            data_start_row = 6

        # Student data
        for si, student in enumerate(students):
            row = data_start_row + si
            cls.apply_cell(ws.cell(row=row, column=1), student.get('id', ''),
                          FONT_DATA, ALIGN_CENTER, BORDER_THIN)
            cls.apply_cell(ws.cell(row=row, column=2), student.get('name', ''),
                          FONT_DATA, ALIGN_CENTER, BORDER_THIN)
            cls.apply_cell(ws.cell(row=row, column=3), student.get('gender', ''),
                          FONT_DATA, ALIGN_CENTER, BORDER_THIN)
            for di in range(num_dims):
                cls.apply_cell(ws.cell(row=row, column=4 + di),
                              None, FONT_DATA, ALIGN_CENTER, BORDER_THIN)
            cls.apply_cell(ws.cell(row=row, column=score_col_idx),
                          None, FONT_DATA, ALIGN_CENTER, BORDER_THIN)
            ws.merge_cells(start_row=row, start_column=remark_col_idx,
                          end_row=row, end_column=remark_col_idx + 1)
            cls.apply_cell(ws.cell(row=row, column=remark_col_idx),
                          None, FONT_DATA, ALIGN_CENTER, BORDER_THIN)

        # Group section
        if group_dims:
            group_start_row = data_start_row + len(students) + 2
            cls._build_group_section(ws, group_start_row, total_cols,
                                    group_dims, num_groups)

    @classmethod
    def _build_info_row(cls, ws, school_name, class_name, student_count,
                        total_cols, score_col_idx):
        cls.apply_cell(ws['A3'], '学校', FONT_INFO_LABEL, ALIGN_CENTER)
        cls.merge_and_write(ws, 'B3:C3', school_name or '', FONT_INFO_VALUE,
                           ALIGN_CENTER, BORDER_BOTTOM_MEDIUM)
        class_label_col = score_col_idx - 3
        class_val_start = class_label_col + 1
        class_val_end = class_val_start + 1
        count_label_col = class_val_end + 1
        count_val_col = count_label_col + 1
        if class_label_col > 3:
            cls.apply_cell(ws.cell(row=3, column=class_label_col),
                          '班级', FONT_INFO_LABEL, ALIGN_CENTER)
            cls.merge_and_write(ws,
                f'{get_column_letter(class_val_start)}3:{get_column_letter(class_val_end)}3',
                class_name or '', FONT_INFO_VALUE, ALIGN_CENTER, BORDER_BOTTOM_MEDIUM)
        if count_label_col < score_col_idx:
            cls.apply_cell(ws.cell(row=3, column=count_label_col),
                          '人数', FONT_INFO_LABEL, ALIGN_CENTER)
            cls.apply_cell(ws.cell(row=3, column=count_val_col),
                          student_count, FONT_INFO_VALUE, ALIGN_CENTER, BORDER_BOTTOM_MEDIUM)

    @classmethod
    def _build_single_row_header(cls, ws, columns, dim_cols,
                                 score_col_name, remark_col_name, total_cols):
        header_row = 5
        for i, col in enumerate(columns):
            cls.apply_cell(ws.cell(row=header_row, column=i + 1),
                          col['header'], FONT_HEADER, ALIGN_CENTER, BORDER_THIN)
        score_col_idx = len(columns) + 1
        remark_col_idx = score_col_idx + 1
        cls.apply_cell(ws.cell(row=header_row, column=score_col_idx),
                      score_col_name, FONT_HEADER, ALIGN_CENTER, BORDER_THIN)
        cls.merge_and_write(ws,
            f'{get_column_letter(remark_col_idx)}{header_row}:{get_column_letter(remark_col_idx + 1)}{header_row}',
            remark_col_name, FONT_HEADER, ALIGN_CENTER, BORDER_THIN)

    @classmethod
    def _build_multi_row_header(cls, ws, columns, dim_cols, score_col_name,
                                remark_col_name, total_cols, has_teacher, has_rule):
        header_row, teacher_row, rule_row = 5, 6, 7
        last_header_row = rule_row if has_rule else teacher_row
        ident_cols = columns[:3]
        score_col_idx = len(columns) + 1
        remark_col_idx = score_col_idx + 1

        for i, col in enumerate(columns):
            col_idx = i + 1
            cls.apply_cell(ws.cell(row=header_row, column=col_idx),
                          col['header'], FONT_HEADER, ALIGN_CENTER, BORDER_THIN)
            if has_teacher:
                cls.apply_cell(ws.cell(row=teacher_row, column=col_idx),
                              col.get('teacher', ''), FONT_DATA, ALIGN_CENTER, BORDER_THIN)
            if has_rule:
                cls.apply_cell(ws.cell(row=rule_row, column=col_idx),
                              col.get('rule', ''), FONT_DATA, ALIGN_CENTER, BORDER_THIN)

        cls.apply_cell(ws.cell(row=header_row, column=score_col_idx),
                      score_col_name, FONT_HEADER, ALIGN_CENTER, BORDER_THIN)
        cls.apply_cell(ws.cell(row=header_row, column=remark_col_idx),
                      remark_col_name, FONT_HEADER, ALIGN_CENTER, BORDER_THIN)

        for i in range(len(ident_cols)):
            ws.merge_cells(start_row=header_row, start_column=i + 1,
                          end_row=last_header_row, end_column=i + 1)
        ws.merge_cells(start_row=header_row, start_column=score_col_idx,
                      end_row=last_header_row, end_column=score_col_idx)
        ws.merge_cells(start_row=header_row, start_column=remark_col_idx,
                      end_row=last_header_row, end_column=remark_col_idx + 1)

        for i in range(1, total_cols + 1):
            for r in range(header_row, last_header_row + 1):
                cls.apply_cell(ws.cell(row=r, column=i), border=BORDER_THIN)

    @classmethod
    def _build_group_section(cls, ws, start_row, total_cols, group_dims, num_groups):
        num_gdims = len(group_dims)
        group_total_col = 2 + num_gdims

        ws.merge_cells(start_row=start_row, start_column=1,
                      end_row=start_row, end_column=total_cols)
        cls.apply_cell(ws.cell(row=start_row, column=1), '小组分数',
                      FONT_HEADER, ALIGN_CENTER)

        sr = start_row + 1
        cls.apply_cell(ws.cell(row=sr, column=1), '组号',
                      FONT_HEADER, ALIGN_CENTER, BORDER_THIN)
        for di, gd in enumerate(group_dims):
            cls.apply_cell(ws.cell(row=sr, column=2 + di),
                          gd['header'], FONT_HEADER, ALIGN_CENTER, BORDER_THIN)
        cls.apply_cell(ws.cell(row=sr, column=group_total_col),
                      '总分', FONT_HEADER, ALIGN_CENTER, BORDER_THIN)

        rr = sr + 1
        cls.apply_cell(ws.cell(row=rr, column=1), '', FONT_DATA, ALIGN_CENTER, BORDER_THIN)
        for di, gd in enumerate(group_dims):
            cls.apply_cell(ws.cell(row=rr, column=2 + di),
                          gd.get('rule', ''), FONT_DATA, ALIGN_CENTER, BORDER_THIN)
        cls.apply_cell(ws.cell(row=rr, column=group_total_col),
                      '', FONT_DATA, ALIGN_CENTER, BORDER_THIN)

        for gi in range(num_groups):
            row = rr + 1 + gi
            cls.apply_cell(ws.cell(row=row, column=1),
                          gi + 1, FONT_DATA, ALIGN_CENTER, BORDER_THIN)
            for di in range(num_gdims):
                cls.apply_cell(ws.cell(row=row, column=2 + di),
                              None, FONT_DATA, ALIGN_CENTER, BORDER_THIN)
            sum_start = get_column_letter(2)
            sum_end = get_column_letter(2 + num_gdims - 1)
            cls.apply_cell(ws.cell(row=row, column=group_total_col),
                          f'=SUM({sum_start}{row}:{sum_end}{row})',
                          FONT_DATA, ALIGN_CENTER, BORDER_THIN)
