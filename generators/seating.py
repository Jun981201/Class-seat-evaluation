"""Seating chart generator — matches original template layout exactly.

Template format:
- Col A: 2.375 (spacer), Col B-C: 15.0 each (left desk pair)
- Col D: 3.25 (aisle), Col E-F: 15.0 each (right desk pair)
- Col G: 4.5 (door), Col H: 4.75 (duty label), Col I: 10.625 (duty values)
- Row 1: Title "3D造物座位表" (merged C1:E1), height=24
- Row 2: School/Class info, height=27
- Row 3: spacer, height=14
- Row 4: header row, height varies
- Rows 5-12: student desk area
- Row 14: 讲台 (merged C14:E14), height=39
- Row 15: empty (no legend)
"""
import random
from openpyxl.styles import Font, Border
import openpyxl
from .template_engine import (
    TemplateEngine, ALIGN_CENTER, BORDER_THIN, BORDER_MEDIUM
)


# ── Exact template fonts ──
FONT_SEAT_TITLE = Font(name='宋体', size=18, bold=True)
FONT_SEAT_INFO = Font(name='宋体', size=11)
FONT_SEAT_STUDENT = Font(name='宋体', size=14)
FONT_DUTY_HEADER = Font(name='宋体', size=11, bold=True)
FONT_DUTY_DAY = Font(name='宋体', size=11, bold=True)
FONT_DUTY_VAL = Font(name='宋体', size=11)

# ── Exact template column widths ──
SEAT_COL_WIDTHS = {
    'A': 2.375, 'B': 15.0, 'C': 15.0, 'D': 3.25,
    'E': 15.0, 'F': 15.0, 'G': 4.5, 'H': 4.75, 'I': 10.625
}

# ── Exact template row heights ──
SEAT_ROW_HEIGHTS = {
    1: 24.0, 2: 27.0, 3: 14.0, 4: 61.0,
    5: 61.0, 6: 61.0, 7: 61.0, 8: 61.0,
    9: 61.0, 10: 61.0, 11: 61.0, 12: 61.0,
    13: 14.0, 14: 39.0, 15: 18.0,
}

# ── Gender fills ──
from openpyxl.styles import PatternFill
FILL_FEMALE = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
FILL_MALE = PatternFill(fill_type=None)  # no fill = white


class SeatingGenerator(TemplateEngine):
    """Generate seating chart matching original template format."""

    MAX_SEATS = 28

    @staticmethod
    def _shuffle_students(students, mode='random', separate_gender=True):
        if not students:
            return []
        males = [s for s in students if s.get('gender', '男') == '男']
        females = [s for s in students if s.get('gender', '男') == '女']

        if mode == 'gender_alternate':
            # Ensure adjacent pairs (B-C, E-F at same row) have different genders
            random.shuffle(males)
            random.shuffle(females)
            result = []
            mi, fi = 0, 0
            while len(result) < len(students):
                # Alternate pair lead: M,F then F,M to avoid same-gender clusters
                if (len(result) // 2) % 2 == 0:
                    if mi < len(males):
                        result.append(males[mi]); mi += 1
                    elif fi < len(females):
                        result.append(females[fi]); fi += 1
                    if fi < len(females):
                        result.append(females[fi]); fi += 1
                    elif mi < len(males):
                        result.append(males[mi]); mi += 1
                else:
                    if fi < len(females):
                        result.append(females[fi]); fi += 1
                    elif mi < len(males):
                        result.append(males[mi]); mi += 1
                    if mi < len(males):
                        result.append(males[mi]); mi += 1
                    elif fi < len(females):
                        result.append(females[fi]); fi += 1
            return result

        elif mode == 'same_gender':
            # Each block of 3 adjacent seats (same desk side) shares the same gender
            random.shuffle(males)
            random.shuffle(females)
            result = []
            mi, fi = 0, 0
            gender_toggle = random.random() < 0.5
            while len(result) < len(students):
                block_size = min(3, len(students) - len(result))
                if gender_toggle:
                    for _ in range(block_size):
                        if mi < len(males):
                            result.append(males[mi]); mi += 1
                        elif fi < len(females):
                            result.append(females[fi]); fi += 1
                else:
                    for _ in range(block_size):
                        if fi < len(females):
                            result.append(females[fi]); fi += 1
                        elif mi < len(males):
                            result.append(males[mi]); mi += 1
                gender_toggle = not gender_toggle
            return result

        else:
            result = list(students)
            random.shuffle(result)
            return result

    @classmethod
    def generate(cls, students, school_name, class_name,
                 duty_schedule=None, arrange_mode='random', separate_gender=True):
        """Generate a minimal, clean seating chart workbook."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Sheet 1: 座位分配 (data reference)
        ws_data = wb.create_sheet('座位分配')
        cls._build_data_sheet(ws_data, students, arrange_mode, separate_gender)

        # Sheet 2: 座位图 (visual layout)
        ws_chart = wb.create_sheet('座位图')
        cls._build_clean_chart(ws_chart, students, school_name, class_name,
                               arrange_mode, separate_gender, duty_schedule)

        return wb

    @classmethod
    def _build_data_sheet(cls, ws, students, arrange_mode='random', separate_gender=True):
        """Simple data reference table."""
        headers = ['座位号', '学号', '姓名', '性别']
        col_widths = {'A': 10, 'B': 10, 'C': 12, 'D': 8}
        cls.set_col_widths(ws, col_widths)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(name='宋体', size=11, bold=True)
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN

        arranged = cls._shuffle_students(students, arrange_mode, separate_gender)
        for i, s in enumerate(arranged[:cls.MAX_SEATS]):
            row = i + 2
            cls.apply_cell(ws.cell(row=row, column=1), i + 1, FONT_SEAT_STUDENT, ALIGN_CENTER, BORDER_THIN)
            cls.apply_cell(ws.cell(row=row, column=2), s.get('id', ''), FONT_SEAT_STUDENT, ALIGN_CENTER, BORDER_THIN)
            cls.apply_cell(ws.cell(row=row, column=3), s.get('name', ''), FONT_SEAT_STUDENT, ALIGN_CENTER, BORDER_THIN)
            cls.apply_cell(ws.cell(row=row, column=4), s.get('gender', ''), FONT_SEAT_STUDENT, ALIGN_CENTER, BORDER_THIN)

    @classmethod
    def _build_clean_chart(cls, ws, students, school_name, class_name,
                           arrange_mode, separate_gender, duty_schedule):
        """Build seating chart matching original template layout exactly."""

        # ── Set exact column widths from template ──
        cls.set_col_widths(ws, SEAT_COL_WIDTHS)

        # ── Set exact row heights from template ──
        cls.set_row_heights(ws, SEAT_ROW_HEIGHTS)

        # ── Row 1: Title ──
        ws.merge_cells('C1:E1')
        cls.apply_cell(ws['C1'], '3D造物座位表', FONT_SEAT_TITLE, ALIGN_CENTER)

        # ── Row 2: School / Class info (labels no border, values with border) ──
        BORDER_NONE = Border()
        cls.apply_cell(ws['B2'], '学校', FONT_SEAT_STUDENT, ALIGN_CENTER, BORDER_NONE)
        cls.apply_cell(ws['C2'], school_name, FONT_SEAT_STUDENT, ALIGN_CENTER, BORDER_THIN)
        cls.apply_cell(ws['E2'], '班级', FONT_SEAT_STUDENT, ALIGN_CENTER, BORDER_NONE)
        cls.apply_cell(ws['F2'], class_name, FONT_SEAT_STUDENT, ALIGN_CENTER, BORDER_THIN)

        # ── Row 4: Desk header areas + Duty header ──
        ws.merge_cells('B4:C4')
        ws.merge_cells('E4:F4')
        ws.merge_cells('H4:I4')
        cls.apply_cell(ws['H4'], '值日安排\n（学号）', FONT_DUTY_HEADER, ALIGN_CENTER)

        # ── Arrange students ──
        arranged = cls._shuffle_students(students, arrange_mode, separate_gender)

        # Grid mapping: seats grouped by desk side (3 adjacent per side).
        # ≤24 students: 6 per desk (no end seat). >24: 7 per desk (with end seat).
        # Always fill front desks first (closest to podium at row 14).
        # Desk 3 (左前 B-C, rows 10-12) → Desk 4 (右前 E-F) → Desk 1 (左后) → Desk 2 (右后)
        total = min(len(arranged), cls.MAX_SEATS)
        use_end_seat = total > 24

        def _desk_seats(col1, col2, rows, end=None):
            """Build seats for one desk: side1(3) + side2(3) [+ end(1)]."""
            seats = []
            for r in rows:
                seats.append((r, col1))
            for r in rows:
                seats.append((r, col2))
            if end:
                seats.append(end)
            return seats

        seat_grid = []
        seat_grid += _desk_seats('B', 'C', [10, 11, 12], (9, 'B') if use_end_seat else None)
        seat_grid += _desk_seats('E', 'F', [10, 11, 12], (9, 'E') if use_end_seat else None)
        seat_grid += _desk_seats('B', 'C', [5, 6, 7],    (8, 'B') if use_end_seat else None)
        seat_grid += _desk_seats('E', 'F', [5, 6, 7],    (8, 'E') if use_end_seat else None)

        # Place students
        for idx, s in enumerate(arranged[:len(seat_grid)]):
            if idx >= len(seat_grid):
                break
            row, col_letter = seat_grid[idx]
            col_num = ord(col_letter) - ord('A') + 1
            display = f"{s.get('id', '')} {s.get('name', '')}"
            fill = FILL_FEMALE if s.get('gender') == '女' else FILL_MALE
            cls.apply_cell(ws.cell(row=row, column=col_num),
                          display, FONT_SEAT_STUDENT, ALIGN_CENTER, BORDER_THIN, fill)

        # ── Row 9 merges (from template) ──
        ws.merge_cells('B9:C9')
        ws.merge_cells('E9:F9')

        # ── Row 14: Podium ──
        ws.merge_cells('C14:E14')
        cls.apply_cell(ws['C14'], '讲  台', FONT_SEAT_TITLE, ALIGN_CENTER)

        # ── No legend / no extra text on row 15 ──
        # (Row 15 is left empty)

        # ── Duty schedule (rows 5-9, cols H-I) ──
        weekdays = ['一', '二', '三', '四', '五']
        for wi, wd in enumerate(weekdays):
            row = 5 + wi
            cls.apply_cell(ws.cell(row=row, column=8), wd,
                          FONT_DUTY_DAY, ALIGN_CENTER, BORDER_THIN)
            if duty_schedule and wd in duty_schedule:
                ids_str = '、'.join(str(i) for i in duty_schedule[wd])
                cls.apply_cell(ws.cell(row=row, column=9),
                              ids_str, FONT_DUTY_VAL, ALIGN_CENTER, BORDER_THIN)
            else:
                cls.apply_cell(ws.cell(row=row, column=9),
                              '', FONT_DUTY_VAL, ALIGN_CENTER, BORDER_THIN)

    @classmethod
    def _assign_positions(cls, arranged_students):
        """Return list of position dicts for preview API."""
        positions = []
        total = min(len(arranged_students), cls.MAX_SEATS)
        use_end_seat = total > 24

        def _desk_seats(col1, col2, rows, end=None):
            seats = []
            for r in rows:
                seats.append((r, col1))
            for r in rows:
                seats.append((r, col2))
            if end:
                seats.append(end)
            return seats

        seat_grid = []
        seat_grid += _desk_seats('B', 'C', [10, 11, 12], (9, 'B') if use_end_seat else None)
        seat_grid += _desk_seats('E', 'F', [10, 11, 12], (9, 'E') if use_end_seat else None)
        seat_grid += _desk_seats('B', 'C', [5, 6, 7],    (8, 'B') if use_end_seat else None)
        seat_grid += _desk_seats('E', 'F', [5, 6, 7],    (8, 'E') if use_end_seat else None)

        for idx, s in enumerate(arranged_students[:len(seat_grid)]):
            row, col_letter = seat_grid[idx]
            # Determine desk/group from position
            if col_letter in ('B', 'C'):
                side = '左'
            else:
                side = '右'
            if row <= 8:
                area = '后'
            else:
                area = '前'

            positions.append({
                'student': s,
                'row': row,
                'col': col_letter,
                'desk_name': f'{area}{side}',
            })

        return positions
